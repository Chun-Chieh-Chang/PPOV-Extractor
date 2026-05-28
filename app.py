import os
import sys
import io
import json
import webbrowser
from threading import Timer
import hashlib
from flask import Flask, jsonify, request, render_template, send_file, session
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ?–å??‰ç”¨ç¨‹å??„åŸºç¤è·¯å¾‘ï??¯æ´ PyInstaller ?“å?ï¼?

def get_resource_path():
    if getattr(sys, \"frozen\", False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def get_data_dir():
    if getattr(sys, \"frozen\", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

RESOURCE_PATH = get_resource_path()
DATA_DIR = get_data_dir()


# Ensure workspace is in python path
sys.path.append(RESOURCE_PATH)
from main import extract_data_from_pdf, _select_directory_dialog, _save_file_dialog, _select_files_dialog

app = Flask(__name__, 
            static_folder=os.path.join(RESOURCE_PATH, 'static'),
            template_folder=os.path.join(RESOURCE_PATH, 'templates'))

app.secret_key = "ppov_extractor_secret_key_123!"

def load_users():
    users_path = os.path.join(DATA_DIR, "users.json")
    if not os.path.exists(users_path):
        try:
            default_data = {
                "users": [
                    {
                        "username": "admin",
                        "role": "admin",
                        "display_name": "ç³»çµ±ç®¡ç???,
                        "password_hash": "3b612c75a7b5048a435fb6ec81e52ff92d6d795a8b5a9c17070f6a63c97a53b2"
                    }
                ]
            }
            with open(users_path, "w", encoding="utf-8") as f:
                json.dump(default_data, f, ensure_ascii=False, indent=2)
            print("Successfully initialized default users.json")
        except Exception as e:
            print(f"Error initializing default users.json: {e}")

    if os.path.exists(users_path):
        try:
            with open(users_path, "r", encoding="utf-8") as f:
                return json.load(f).get("users", [])
        except Exception as e:
            print(f"Error loading users.json: {e}")
    return []

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify({"success": False, "message": "?’ç?å­˜å?ï¼šæ‚¨ä¸å…·?™ç®¡?†å“¡æ¬Šé?ï¼?}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    payload = request.json or {}
    username = payload.get("username", "").strip().lower()
    password = payload.get("password", "")
    
    if not username or not password:
        return jsonify({"success": False, "message": "è«‹è¼¸?¥å¸³?Ÿè?å¯†ç¢¼"})
        
    password_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
    users = load_users()
    
    user = next((u for u in users if u["username"].lower() == username and u["password_hash"] == password_hash), None)
    
    if not user:
        return jsonify({"success": False, "message": "å¸³è??–å?ç¢¼éŒ¯èª?})
        
    session["username"] = user["username"]
    session["role"] = user["role"]
    session["display_name"] = user["display_name"]
    
    return jsonify({
        "success": True,
        "message": "?»å…¥?å?ï¼?,
        "user": {
            "username": user["username"],
            "role": user["role"],
            "display_name": user["display_name"]
        }
    })

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"success": True, "message": "å·²æ??Ÿç™»??})

@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    if "username" in session:
        return jsonify({
            "success": True,
            "logged_in": True,
            "user": {
                "username": session["username"],
                "role": session["role"],
                "display_name": session["display_name"]
            }
        })
    return jsonify({
        "success": True,
        "logged_in": False,
        "user": {
            "username": "guest",
            "role": "inspector",
            "display_name": "?è³ªæª¢æŸ¥??
        }
    })

@app.route("/api/auth/change_password", methods=["POST"])
@admin_required
def auth_change_password():
    payload = request.json or {}
    current_password = payload.get("current_password", "")
    new_password = payload.get("new_password", "")
    confirm_password = payload.get("confirm_password", "")

    if not current_password or not new_password or not confirm_password:
        return jsonify({"success": False, "message": "è«‹å¡«å¯«æ??‰æ?ä½?})

    if new_password != confirm_password:
        return jsonify({"success": False, "message": "?°å?ç¢¼è?ç¢ºè?å¯†ç¢¼ä¸ä???})

    if len(new_password) < 6:
        return jsonify({"success": False, "message": "?°å?ç¢¼é•·åº¦è‡³å°‘é?è¦?6 ?‹å???})

    username = session.get("username")
    current_hash = hashlib.sha256(current_password.encode("utf-8")).hexdigest()
    new_hash = hashlib.sha256(new_password.encode("utf-8")).hexdigest()

    users_path = os.path.join(DATA_DIR, "users.json")
    try:
        with open(users_path, "r", encoding="utf-8") as f:
            users_data = json.load(f)

        user_found = False
        for user in users_data.get("users", []):
            if user["username"].lower() == username.lower():
                if user["password_hash"] != current_hash:
                    return jsonify({"success": False, "message": "?®å?å¯†ç¢¼?¯èª¤ï¼Œè??æ–°ç¢ºè?"})
                user["password_hash"] = new_hash
                user_found = True
                break

        if not user_found:
            return jsonify({"success": False, "message": "?¾ä??°ä½¿?¨è€…è???})

        with open(users_path, "w", encoding="utf-8") as f:
            json.dump(users_data, f, ensure_ascii=False, indent=2)

        return jsonify({"success": True, "message": "å¯†ç¢¼å·²æ??Ÿæ›´?°ï?"})

    except Exception as e:
        print(f"Error changing password: {e}")
        return jsonify({"success": False, "message": f"?²å?å¤±æ?ï¼š{str(e)}"})


db = {
    "extracted_data": [],
    "config": None,
    "last_folder": ""
}

def get_db_file_path():
    return os.path.join(DATA_DIR, "ppov_database.json")

def load_db_from_file():
    db_path = get_db_file_path()
    if os.path.exists(db_path):
        try:
            with open(db_path, "r", encoding="utf-8") as f:
                db["extracted_data"] = json.load(f)
            print(f"Loaded {len(db['extracted_data'])} records from ppov_database.json")
        except Exception as e:
            print(f"Error loading ppov_database.json: {e}")
            db["extracted_data"] = []
    else:
        db["extracted_data"] = []

def save_db_to_file():
    db_path = get_db_file_path()
    try:
        with open(db_path, "w", encoding="utf-8") as f:
            json.dump(db["extracted_data"], f, ensure_ascii=False, indent=2)
        print(f"Successfully saved {len(db['extracted_data'])} records to ppov_database.json")
    except Exception as e:
        print(f"Error saving ppov_database.json: {e}")

def load_config():
    config_path = os.path.join(DATA_DIR, "config.json")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            db["config"] = json.load(f)
    except Exception as e:
        print(f"Error loading config.json: {e}")
    load_db_from_file()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/favicon.ico")
def favicon():
    """Silences browser favicon.ico 404 console errors by returning 204 No Content."""
    return "", 204

@app.route("/api/config", methods=["GET"])
def get_config_endpoint():
    if not db["config"]:
        load_config()
    return jsonify(db["config"])

@app.route("/api/select_folder", methods=["POST"])
@admin_required
def select_folder():
    """Triggers native OS directory picker."""
    try:
        # Run dialog safely
        selected_path = _select_directory_dialog("?¸æ??…å« PPOV PDF ?„è??™å¤¾", db["last_folder"])
        if selected_path:
            db["last_folder"] = selected_path
            return jsonify({"success": True, "path": selected_path})
        return jsonify({"success": False, "message": "?ªé¸?‡ä»»ä½•è??™å¤¾"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route("/api/db", methods=["GET"])
def get_database():
    if not db["config"]:
        load_config()
    load_db_from_file()
    return jsonify({
        "success": True,
        "count": len(db["extracted_data"]),
        "data": db["extracted_data"],
        "fields": [f["name"] for f in db["config"]["fields_to_extract"]] if db["config"] else []
    })

@app.route("/api/db/add", methods=["POST"])
@admin_required
def db_add_record():
    new_record = request.json
    if not new_record or not new_record.get("?¢å??‹è?"):
        return jsonify({"success": False, "message": "?¢å??‹è?ï¼ˆå??Ÿï??ºå?å¡«é?"})
    
    part_no = new_record.get("?¢å??‹è?").strip()
    
    # Check duplicate
    if any(item.get("?¢å??‹è?") == part_no for item in db["extracted_data"]):
        return jsonify({"success": False, "message": f"?è? {part_no} å·²å??¨æ–¼è³‡æ?åº«ä¸­"})
    
    # Supply defaults
    if not new_record.get("æª”æ??ç¨±"):
        new_record["æª”æ??ç¨±"] = f"MANUAL_{part_no}.pdf"
        
    db["extracted_data"].append(new_record)
    save_db_to_file()
    
    return jsonify({
        "success": True, 
        "message": f"?è? {part_no} ?°å??å?", 
        "data": db["extracted_data"]
    })

@app.route("/api/db/edit", methods=["POST"])
@admin_required
def db_edit_record():
    edit_data = request.json
    if not edit_data or not edit_data.get("?¢å??‹è?"):
        return jsonify({"success": False, "message": "?¡æ??„ä¿®?¹è?æ±‚ï??è?å¿…å¡«"})
        
    part_no = edit_data.get("?¢å??‹è?").strip()
    
    # Find and update
    found = False
    for i, item in enumerate(db["extracted_data"]):
        if item.get("?¢å??‹è?") == part_no:
            # Update values
            for k, v in edit_data.items():
                item[k] = v
            found = True
            break
            
    if not found:
        return jsonify({"success": False, "message": f"?¨è??™åº«ä¸­æ‰¾ä¸åˆ°?è? {part_no}"})
        
    save_db_to_file()
    return jsonify({
        "success": True, 
        "message": f"?è? {part_no} ä¿®æ”¹?å?", 
        "data": db["extracted_data"]
    })

@app.route("/api/db/delete", methods=["POST"])
@admin_required
def db_delete_record():
    payload = request.json or {}
    part_no = payload.get("part_no")
    if not part_no:
        return jsonify({"success": False, "message": "?¡æ??„åˆª?¤è?æ±‚ï??è?å¿…å¡«"})
        
    initial_len = len(db["extracted_data"])
    db["extracted_data"] = [item for item in db["extracted_data"] if item.get("?¢å??‹è?") != part_no]
    
    if len(db["extracted_data"]) == initial_len:
        return jsonify({"success": False, "message": f"?¨è??™åº«ä¸­æ‰¾ä¸åˆ°?è? {part_no}"})
        
    save_db_to_file()
    return jsonify({
        "success": True, 
        "message": f"?è? {part_no} ?ªé™¤?å?", 
        "data": db["extracted_data"]
    })

@app.route("/api/db/clear", methods=["POST"])
@admin_required
def db_clear():
    db["extracted_data"] = []
    save_db_to_file()
    return jsonify({
        "success": True, 
        "message": "è³‡æ?åº«å·²å®Œå…¨æ¸…ç©º", 
        "data": []
    })

@app.route("/api/db/import_pdf", methods=["POST"])
@admin_required
def db_import_single_pdf():
    """?¥æ”¶?®ä? PPOV PDF æª”æ?ï¼Œæ??–å…¶?å??ƒæ•¸ï¼Œä¸¦?ªå??°å?/?´æ–°?³è??™åº«ä¸­ã€?""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "?ªæ”¶?°æ?æ¡?})
            
        f = request.files["file"]
        if not f or f.filename == "":
            return jsonify({"success": False, "message": "?ªé¸?‡ä»»ä½•æ?æ¡?})
            
        if not f.filename.lower().endswith(".pdf"):
            return jsonify({"success": False, "message": "ä¸æ”¯?´ç??¼å?ï¼Œè??¸æ? .pdf è¦æ ¼?®æ?æ¡?})
            
        filename = f.filename
        
        # ?«å???output/ ?®é?ä¸­ä»¥ä¾¿é€²è?å¯¦é?è·¯å??å?
        temp_dir = os.path.join(DATA_DIR, "output")
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            
        temp_filepath = os.path.join(temp_dir, filename)
        f.save(temp_filepath)
        
        if not db["config"]:
            load_config()
            
        # èª¿ç”¨?¸å??å??½å?
        data = extract_data_from_pdf(temp_filepath, db["config"])
        
        # ç§»é™¤?«å?æª”æ?
        if os.path.exists(temp_filepath):
            os.remove(temp_filepath)
            
        if not data:
            return jsonify({"success": False, "message": f"æª”æ? {filename} è§??å¤±æ??–æ ¼å¼ä???})
            
        part_no = data.get("?¢å??‹è?", "").strip()
        if not part_no or part_no == "?ªæ‰¾??:
            return jsonify({"success": False, "message": f"??PDF è¦æ ¼??{filename} ä¸­æœª?¾åˆ°?‰æ??„ç”¢?å???})
            
        load_db_from_file()
        
        # æª¢æŸ¥?¯å¦å·²å??¨ï??²è?è¦†è?/?°å??ˆä½µ
        existing_idx = next((i for i, item in enumerate(db["extracted_data"]) if item.get("?¢å??‹è?") == part_no), None)
        
        if existing_idx is not None:
            db["extracted_data"][existing_idx] = data
            msg = f"?è? {part_no} å·²å??¨æ–¼è³‡æ?åº«ä¸­ï¼Œå·²?å??æ–°è§?? PDF ä¸¦è??‹è??¼å??¸ï?"
        else:
            db["extracted_data"].append(data)
            msg = f"?è? {part_no} è§???å?ï¼Œå·²?ªå?å°å…¥è¦æ ¼è³‡æ?åº«ï?"
            
        save_db_to_file()
        
        return jsonify({
            "success": True,
            "message": msg,
            "count": len(db["extracted_data"]),
            "data": db["extracted_data"],
            "fields": [f["name"] for f in db["config"]["fields_to_extract"]] if db["config"] else []
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route("/api/db/import_pdf_native", methods=["POST"])
@admin_required
def db_import_pdf_native():
    """?é?å¾Œç«¯?Ÿç? OS æª”æ??¸æ?è¦–ç?å°å…¥ä¸€?–å???PPOV PDF æª”æ?ä¸¦è‡ª?•è§£?ã€å?æª”ã€?""
    try:
        # é¡¯ç¤º?Ÿç?æª”æ?å¤šé¸å½ˆç?
        selected_files = _select_files_dialog(
            title="?¸æ? PPOV PDF è¦æ ¼?¸æ?æ¡?(?¯å???",
            initial_dir=db["last_folder"] or DATA_DIR,
            file_types=[("PDF Files", "*.pdf")]
        )
        
        if not selected_files:
            return jsonify({"success": False, "message": "?ªé¸?‡ä»»ä½?PDF æª”æ?"})
            
        if not db["config"]:
            load_config()
            
        load_db_from_file()
        
        imported_parts = []
        updated_count = 0
        added_count = 0
        
        for filepath in selected_files:
            if not os.path.exists(filepath):
                continue
            
            # ?´æ–° last_folder ?ºæ?å¾Œé¸?‡æ?æ¡ˆç?è³‡æ?å¤?
            db["last_folder"] = os.path.dirname(filepath)
            
            # èª¿ç”¨?¸å??å??½å?
            data = extract_data_from_pdf(filepath, db["config"])
            if not data:
                continue
                
            part_no = data.get("?¢å??‹è?", "").strip()
            if not part_no or part_no == "?ªæ‰¾??:
                continue
                
            # æª¢æŸ¥?¯å¦å·²å??¨ï??²è?è¦†è?/?°å??ˆä½µ
            existing_idx = next((i for i, item in enumerate(db["extracted_data"]) if item.get("?¢å??‹è?") == part_no), None)
            
            if existing_idx is not None:
                db["extracted_data"][existing_idx] = data
                updated_count += 1
            else:
                db["extracted_data"].append(data)
                added_count += 1
                
            imported_parts.append(part_no)
            
        if not imported_parts:
            return jsonify({"success": False, "message": "?€?¸ç?æª”æ??†è§£?å¤±?—æ??¼å?ä¸ç¬¦"})
            
        save_db_to_file()
        
        # å»ºç??å ±è¨Šæ¯
        msg = f"?å?å°å…¥ {len(imported_parts)} ç­†è??¼ï?"
        if added_count > 0:
            msg += f" ?°å? {added_count} ç­?
        if updated_count > 0:
            msg += f" è¦†è?/?´æ–° {updated_count} ç­?
            
        # ?¥åªå°å…¥?®ä??è?ï¼Œå??³è©²?è?ä»¥ä¾¿?ç«¯é«˜äº®?‡é?è¦?
        last_part_no = imported_parts[-1] if len(imported_parts) == 1 else None
        
        return jsonify({
            "success": True,
            "message": msg,
            "count": len(db["extracted_data"]),
            "data": db["extracted_data"],
            "last_part_no": last_part_no,
            "fields": [f["name"] for f in db["config"]["fields_to_extract"]] if db["config"] else []
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route("/api/extract", methods=["POST"])
@admin_required
def extract_data():
    """Performs incremental extraction on PDF files in the selected folder."""
    data_payload = request.json or {}
    folder_path = data_payload.get("path", db["last_folder"])
    is_incremental = data_payload.get("incremental", True)
    
    if not folder_path or not os.path.exists(folder_path):
        return jsonify({"success": False, "message": "?¡æ??„è??™å¤¾è·¯å?"})
    
    if not db["config"]:
        load_config()
    
    # Ensure DB is loaded
    if not db["extracted_data"]:
        load_db_from_file()
        
    pdf_files = []
    for root, dirs, files in os.walk(folder_path):
        dirs[:] = [d for d in dirs if not d.startswith('$') and not d.startswith('.') and d not in ['System Volume Information', 'RECYCLE.BIN']]
        for f in files:
            if f.lower().endswith('.pdf'):
                pdf_files.append(os.path.join(root, f))
    if not pdf_files:
        return jsonify({"success": False, "message": "æ­¤è??™å¤¾?§ç„¡ä»»ä? PDF æª”æ?"})
        
    # Incremental sync: filter out already processed PDF files
    existing_filenames = {item.get("æª”æ??ç¨±") for item in db["extracted_data"] if item.get("æª”æ??ç¨±")}
    
    if is_incremental:
        files_to_process = [p for p in pdf_files if os.path.basename(p) not in existing_filenames]
    else:
        files_to_process = pdf_files
        
    if not files_to_process:
        return jsonify({
            "success": True, 
            "message": "?€??PDF æª”æ??†å·²?¨è??™åº«ä¸­ï??¡é??Œæ­¥ï¼?,
            "count": len(db["extracted_data"]), 
            "data": db["extracted_data"],
            "fields": [f["name"] for f in db["config"]["fields_to_extract"]]
        })
        
    new_results = []
    for pdf_path in files_to_process:
        try:
            data = extract_data_from_pdf(pdf_path, db["config"])
            if data:
                new_results.append(data)
        except Exception as e:
            print(f"Error processing {pdf_path}: {e}")
            
    if is_incremental:
        existing_by_file = {item.get("æª”æ??ç¨±"): item for item in db["extracted_data"] if item.get("æª”æ??ç¨±")}
        for item in new_results:
            existing_by_file[item.get("æª”æ??ç¨±")] = item
        db["extracted_data"] = list(existing_by_file.values())
    else:
        db["extracted_data"] = new_results
        
    save_db_to_file()
    
    return jsonify({
        "success": True, 
        "count": len(db["extracted_data"]), 
        "data": db["extracted_data"],
        "fields": [f["name"] for f in db["config"]["fields_to_extract"]]
    })

@app.route("/api/export_master", methods=["POST"])
@admin_required
def export_master():
    """Generates and exports the master Excel or JSON file in memory."""       
    format_type = request.json.get("format", "excel")
    if not db["extracted_data"]:
        return jsonify({"success": False, "message": "?®å??¡ä»»ä½•å·²?å?ä¹‹æ•¸??}) 
        
    df = pd.DataFrame(db["extracted_data"])
    column_order = ["æª”æ??ç¨±"] + [field["name"] for field in db["config"]["fields_to_extract"]]
    df = df[column_order]

    # Save a backup copy to the configured public folder on the server
    if not db["config"]:
        load_config()
    public_folder = db["config"].get("public_export_folder", "output/public")
    abs_public_folder = os.path.abspath(os.path.join(DATA_DIR, public_folder))
    
    try:
        if not os.path.exists(abs_public_folder):
            os.makedirs(abs_public_folder)
        
        backup_filename = "PPOV_Master_Table.xlsx" if format_type == "excel" else "PPOV_Master_Table.json"
        backup_path = os.path.join(abs_public_folder, backup_filename)
        
        if format_type == "excel":
            df.to_excel(backup_path, index=False, engine='openpyxl')
        else:
            with open(backup_path, "w", encoding="utf-8") as f:
                json.dump(db["extracted_data"], f, ensure_ascii=False, indent=2)
        print(f"Successfully saved server-side public backup to: {backup_path}")
    except Exception as e:
        print(f"Error saving public backup copy: {e}")

    if format_type == "excel":
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name="PPOV_Master_Table.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        content = json.dumps(db["extracted_data"], ensure_ascii=False, indent=2)
        buffer = io.BytesIO(content.encode("utf-8"))
        return send_file(
            buffer,
            as_attachment=True,
            download_name="PPOV_Master_Table.json",
            mimetype="application/json"
        )

@app.route("/api/export_part", methods=["POST"])
def export_part_excel():
    """Generates a highly premium structured Excel sheet for a single part.""" 
    part_no = request.json.get("part_no")
    inspection_data = request.json.get("inspection_data", {})
    if not part_no:
        return jsonify({"success": False, "message": "è«‹æ?å®šå???})
        
    part_data = next((item for item in db["extracted_data"] if item.get("?¢å??‹è?") == part_no), None)
    if not part_data:
        return jsonify({"success": False, "message": f"?¾ä??°å???{part_no} ?„æ•¸??})
        
    # Generate beautifully styled spreadsheet using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"PPOV - {part_no}"
    ws.views.sheetView[0].showGridLines = True
    
    # ?€?€?€ COLOR SYSTEM (Coordinated Ice Blue Light Theme) ?€?€?€
    NAVY_FILL = PatternFill(start_color="1A3A5F", end_color="1A3A5F", fill_type="solid") # Deep Navy Blue
    HEADER_FILL = PatternFill(start_color="3A7CA8", end_color="3A7CA8", fill_type="solid") # Steel Blue
    SUBHEADER_FILL = PatternFill(start_color="50718C", end_color="50718C", fill_type="solid") # Slate Blue
    ACCENT_FILL = PatternFill(start_color="F0F7FB", end_color="F0F7FB", fill_type="solid") # Light Ice Blue
    
    # ?€?€?€ FONTS ?€?€?€
    title_font = Font(name="Microsoft JhengHei", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    label_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="1A3A5F") # Navy label text
    value_font = Font(name="Microsoft JhengHei", size=10, color="000000")
    header_col_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='B4D8E7'), # Light Ice Blue Border
        right=Side(style='thin', color='B4D8E7'),
        top=Side(style='thin', color='B4D8E7'),
        bottom=Side(style='thin', color='B4D8E7')
    )
    double_bottom = Border(bottom=Side(style='double', color='1A3A5F'))
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Title Block will be formatted at Row 1
        
    # --- 1. TITLE BLOCK ---
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"PPOV å°„å‡º?å??¸æ??¥æª¢è¡?- {part_no}"
    title_cell.font = title_font
    title_cell.fill = NAVY_FILL
    title_cell.alignment = center_align
    
    # --- 2. BASIC INFORMATION SECTION (Rows 3-7) ---
    ws.merge_cells("A2:E2")
    info_sec = ws["A2"]
    info_sec.value = "  ?ºæœ¬è³‡è? (Basic Information)"
    info_sec.font = section_font
    info_sec.fill = HEADER_FILL
    info_sec.alignment = Alignment(horizontal="left", vertical="center")
    
    basic_fields = [
        ("?¢å??‹è? Part No.", part_data.get("?¢å??‹è?", "N/A"), "?–é¢?ˆæ¬¡ Drawing Rev.", part_data.get("?–é¢?ˆæ¬¡", "N/A")),
        ("?¢å??ç¨± Description", part_data.get("?¢å??ç¨±", "N/A"), "æ¨¡å…·ç·¨è? Mold No.", part_data.get("æ¨¡å…·ç·¨è?", "N/A")),
        ("æ¨¡å…·ç©´æ•¸ Cavitation", part_data.get("æ¨¡å…·ç©´æ•¸", "N/A"), "å°„å‡º?å?æ©Ÿç·¨??Press No.", part_data.get("å°„å‡º?å?æ©Ÿç·¨??, "N/A")),
        ("æ©Ÿå°?¸æ•¸ Press Tonnage", part_data.get("å°„å‡º?å?æ©Ÿå™¸??, "N/A"), "?ºæ¡¿å°ºå¯¸ Screw Dia.", part_data.get("?ºæ¡¿å°ºå¯¸", "N/A")),
        ("?Ÿæ??™è? Material No.", part_data.get("?Ÿæ??™è?", "N/A"), "?˜æ?æ¢ä»¶ Drying Cond.", part_data.get("?˜æ?æ¢ä»¶", "N/A"))
    ]
    
    curr_row = 3
    for f1, v1, f2, v2 in basic_fields:
        ws.cell(row=curr_row, column=1, value=f1).font = label_font
        ws.cell(row=curr_row, column=1).alignment = left_align
        ws.cell(row=curr_row, column=2, value=v1).font = value_font
        ws.cell(row=curr_row, column=2).alignment = left_align
        ws.cell(row=curr_row, column=3, value=f2).font = label_font
        ws.cell(row=curr_row, column=3).alignment = left_align
        ws.merge_cells(start_row=curr_row, start_column=4, end_row=curr_row, end_column=5)
        ws.cell(row=curr_row, column=4, value=v2).font = value_font
        ws.cell(row=curr_row, column=4).alignment = left_align
        
        # Apply border & soft background to label columns
        for c in range(1, 6):
            cell = ws.cell(row=curr_row, column=c)
            cell.border = thin_border
            if c in [1, 3]:
                cell.fill = ACCENT_FILL
        curr_row += 1
        
    curr_row += 1 # Spacing
    
    # --- 3. PROCESS PARAMETERS TABLE (Rows 9-20) ---
    ws.merge_cells(f"A{curr_row}:E{curr_row}")
    proc_sec = ws.cell(row=curr_row, column=1, value="  ?œéµè£½ç??ƒæ•¸ (Key Process Parameters)")
    proc_sec.font = section_font
    proc_sec.fill = HEADER_FILL
    proc_sec.alignment = Alignment(horizontal="left", vertical="center")
    curr_row += 1
    
    # Columns Headers
    headers = ["?ƒæ•¸?…ç›® Parameter", "?®æ???(Target)", "ä¸‹é???(Low)", "ä¸Šé???(High)", "å¯¦é???(Actual)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=curr_row, column=c, value=h)
        cell.font = header_col_font
        cell.fill = SUBHEADER_FILL
        cell.alignment = center_align
        cell.border = thin_border
    curr_row += 1
    
    proc_rows = [
        ("å¡«å??‚é? Fill Time (s)", "å¡«å??‚é?_?®æ???, "å¡«å??‚é?_ä¸‹é???, "å¡«å??‚é?_ä¸Šé???, "å¡«å??‚é?_å¯¦é???),
        ("ä¿å?å£“å? Hold Pressure (bar)", "ä¿å?å£“å?_?®æ???, "ä¿å?å£“å?_ä¸‹é???, "ä¿å?å£“å?_ä¸Šé???, "ä¿å?å£“å?_å¯¦é???),
        ("ä¿å??‚é? Hold Time (s)", "ä¿å??‚é?_?®æ???, "ä¿å??‚é?_ä¸‹é???, "ä¿å??‚é?_ä¸Šé???, "ä¿å??‚é?_å¯¦é???),
        ("ä¿å?å®Œç”¢?é???Packed Weight (g)", "ä¿å?å®Œç??¢å?å¹³å??é?_?®æ???, "ä¿å?å®Œç??¢å?å¹³å??é?_ä¸‹é???, "ä¿å?å®Œç??¢å?å¹³å??é?_ä¸Šé???, "ä¿å?å®Œç??¢å?å¹³å??é?_å¯¦é???),
        ("?·å»?‚é? Cooling Time (s)", "?·å»?‚é?_?®æ???, "?·å»?‚é?_ä¸‹é???, "?·å»?‚é?_ä¸Šé???, "?·å»?‚é?_å¯¦é???),
        ("æ¨¡å…·æº«åº¦-æ¯æ¨¡ Water Temp A-Side (??", "æ¨¡å…·æº«åº¦è¨­å?-æ¯æ¨¡_?®æ???, "æ¨¡å…·æº«åº¦è¨­å?-æ¯æ¨¡_ä¸‹é???, "æ¨¡å…·æº«åº¦è¨­å?-æ¯æ¨¡_ä¸Šé???, "æ¨¡å…·æº«åº¦è¨­å?-æ¯æ¨¡_å¯¦é???),
        ("æ¨¡å…·æº«åº¦-?¬æ¨¡ Water Temp B-Side (??", "æ¨¡å…·æº«åº¦è¨­å?-?¬æ¨¡_?®æ???, "æ¨¡å…·æº«åº¦è¨­å?-?¬æ¨¡_ä¸‹é???, "æ¨¡å…·æº«åº¦è¨­å?-?¬æ¨¡_ä¸Šé???, "æ¨¡å…·æº«åº¦è¨­å?-?¬æ¨¡_å¯¦é???),
        ("æ¨¡å…·æº«åº¦-æ»‘å? Water Temp Slide (??", "æ¨¡å…·æº«åº¦è¨­å?-æ»‘å?_?®æ???, "æ¨¡å…·æº«åº¦è¨­å?-æ»‘å?_ä¸‹é???, "æ¨¡å…·æº«åº¦è¨­å?-æ»‘å?_ä¸Šé???, "æ¨¡å…·æº«åº¦è¨­å?-æ»‘å?_å¯¦é???),
    ]
    
    for label, target_k, low_k, high_k, actual_k in proc_rows:
        ws.cell(row=curr_row, column=1, value=label).font = label_font
        ws.cell(row=curr_row, column=1).alignment = left_align
        ws.cell(row=curr_row, column=1).fill = ACCENT_FILL
        
        for c, key in enumerate([target_k, low_k, high_k], 2):
            cell = ws.cell(row=curr_row, column=c, value=part_data.get(key, "N/A"))
            cell.font = value_font
            cell.alignment = center_align
            
        # å¯¦é???(Column 5) ?¨éƒ¨?™ç©º
        cell_actual = ws.cell(row=curr_row, column=5, value="")
        cell_actual.font = value_font
        cell_actual.alignment = center_align
            
        for c in range(1, 6):
            ws.cell(row=curr_row, column=c).border = thin_border
        curr_row += 1
        
    curr_row += 1 # Spacing
    
    # --- 4. REFERENCE PARAMETERS SECTION (Rows 22-26) ---
    ws.merge_cells(f"A{curr_row}:E{curr_row}")
    ref_sec = ws.cell(row=curr_row, column=1, value="  ?ƒè€ƒè??¼å???(Reference Parameters)")
    ref_sec.font = section_font
    ref_sec.fill = HEADER_FILL
    ref_sec.alignment = Alignment(horizontal="left", vertical="center")
    curr_row += 1
    
    ref_fields = [
        ("ä¿å?å®Œæ¨¡??Packed Out Shot Weight (g)", "ä¿å?å®Œç?æ¨¡é?_?®æ???),
        ("?–æ¨¡?›è¨­å®?Clamp Tonnage (ton)", "?–æ¨¡?›_?®æ???),
        ("?Ÿç”¢?±æ??‚é? Mold Cycle Time (s)", "?±æ??‚é?_?®æ???)
    ]
    
    for label, key in ref_fields:
        # Column 1: Label
        ws.cell(row=curr_row, column=1, value=label).font = label_font
        ws.cell(row=curr_row, column=1).alignment = left_align
        ws.cell(row=curr_row, column=1).fill = ACCENT_FILL
        
        # Columns 2-4 Merged: Extracted Reference Value
        ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=4)
        ws.cell(row=curr_row, column=2, value=part_data.get(key, "N/A")).font = value_font
        ws.cell(row=curr_row, column=2).alignment = center_align
        
        # Column 5: Blank Check Field (for user inspection)
        ws.cell(row=curr_row, column=5, value="").font = value_font
        ws.cell(row=curr_row, column=5).alignment = center_align
        
        for c in range(1, 6):
            ws.cell(row=curr_row, column=c).border = thin_border
        curr_row += 1
        
    # --- 5. ON-SITE INSPECTION RECORD (Rows 28-30) ---
    curr_row += 1 # Spacing
    ws.merge_cells(f"A{curr_row}:E{curr_row}")
    inspect_sec = ws.cell(row=curr_row, column=1, value="  ?¾å ´?Ÿç”¢?¥æª¢ç´€??(On-site Inspection Record)")
    inspect_sec.font = section_font
    inspect_sec.fill = HEADER_FILL
    inspect_sec.alignment = Alignment(horizontal="left", vertical="center")
    curr_row += 1
    
    # Inspection record fields (2 rows, 4 columns total)
    inspect_fields = [
        ("å¯¦é?æ©Ÿå°ç·¨è? Actual Press No.", "sign_press_no", "?¥æª¢?¥æ? Inspection Date", "sign_date"),
        ("?¥æª¢?‚é? Inspection Time", "sign_time", "?¥æª¢?¡ç°½??Inspector Signature", "sign_inspector")
    ]
    
    for f1, k1, f2, k2 in inspect_fields:
        # Column 1: Label 1
        ws.cell(row=curr_row, column=1, value=f1).font = label_font
        ws.cell(row=curr_row, column=1).alignment = left_align
        ws.cell(row=curr_row, column=1).fill = ACCENT_FILL
        
        # Column 2: Value 1
        ws.cell(row=curr_row, column=2, value=inspection_data.get(k1, "")).font = value_font
        ws.cell(row=curr_row, column=2).alignment = center_align
        
        # Column 3: Label 2
        ws.cell(row=curr_row, column=3, value=f2).font = label_font
        ws.cell(row=curr_row, column=3).alignment = left_align
        ws.cell(row=curr_row, column=3).fill = ACCENT_FILL
        
        # Columns 4-5 Merged: Value 2
        ws.merge_cells(start_row=curr_row, start_column=4, end_row=curr_row, end_column=5)
        ws.cell(row=curr_row, column=4, value=inspection_data.get(k2, "")).font = value_font
        ws.cell(row=curr_row, column=4).alignment = center_align
        
        # Apply borders
        for c in range(1, 6):
            ws.cell(row=curr_row, column=c).border = thin_border
        curr_row += 1
        
    # --- 6. FOOTER SIGNATURE ---
    curr_row += 1 # Spacing row
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
    author_cell = ws.cell(row=curr_row, column=1, value="Wesley Chang @ Mouldex, 2026. QC Dept. | PPOV å°„å‡º?å??¸æ??¥æª¢è¡?)
    author_cell.font = Font(name="Microsoft JhengHei", size=8, italic=True, color="64748B") # Slate 500
    author_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Set optimized print-safe column widths (Total: 78, perfectly fits A4 portrait width)
    ws.column_dimensions['A'].width = 38.5  # Parameter Label
    ws.column_dimensions['B'].width = 18.5  # Target Value
    ws.column_dimensions['C'].width = 30.0  # Low Value
    ws.column_dimensions['D'].width = 11.3  # High Value
    ws.column_dimensions['E'].width = 11.3  # Actual Value/Check Record
    
    # ?€?€?€ ROW HEIGHTS (Only active rows with content, preventing trailing page overflows) ?€?€?€
    ws.row_dimensions[1].height = 40
    for r in range(2, curr_row + 1):
        ws.row_dimensions[r].height = 24
    
    # ?€?€?€ PAGE PRINT SETUP (A4 & Auto Fit to 1 Page Width & Height) ?€?€?€
    ws.page_setup.paperSize = 9  # A4 Paper Size
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f'A1:E{curr_row}'  # Explicitly restrict print area
    
    # Set customized print-safe margins (Left/Right: 1.3cm / 0.51in, Top/Bottom: 0.8cm / 0.31in, Header/Footer: 0)
    ws.page_margins.left = 0.51
    ws.page_margins.right = 0.51
    ws.page_margins.top = 0.31
    ws.page_margins.bottom = 0.31
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    
    # Center on page Horizontally and Vertically
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"PPOV_Spec_{part_no}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/api/load_master_file", methods=["POST"])
def load_master_file():
    """?¥æ”¶?è¦½?¨ä??³ç? Excel ??JSON ç¸½è¡¨æª”æ?ï¼Œå??¨ä?ä¾è³´ tkinter??""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "?ªæ”¶?°æ?æ¡?})

        f = request.files["file"]
        if not f or f.filename == "":
            return jsonify({"success": False, "message": "?ªé¸?‡ä»»ä½•æ?æ¡?})

        filename = f.filename.lower()
        results = []

        if filename.endswith(".xlsx"):
            import io
            file_bytes = io.BytesIO(f.read())
            df = pd.read_excel(file_bytes, engine="openpyxl")
            # ä¸€æ¬¡æ€§æ?æ´—æ???NaN ä¸¦è??ºå?ä¸²ï??¿å??æ ¼?æ­·
            df = df.fillna("")
            for col in df.columns:
                df[col] = df[col].apply(
                    lambda x: "" if x == "" else str(x) if not isinstance(x, str) else x
                )
            results = df.to_dict(orient="records")

        elif filename.endswith(".json"):
            import io
            content = f.read().decode("utf-8")
            results = json.loads(content)

        else:
            return jsonify({"success": False, "message": "ä¸æ”¯?´ç??¼å?ï¼Œè??¸æ? .xlsx ??.json"})

        db["extracted_data"] = results
        save_db_to_file()

        if not db["config"]:
            load_config()

        return jsonify({
            "success": True,
            "count": len(results),
            "data": results,
            "fields": [f["name"] for f in db["config"]["fields_to_extract"]] if db["config"] else []
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})



def launch_browser():
    webbrowser.open("http://127.0.0.1:5000")

if __name__ == "__main__":
    load_config()
    # Start server and auto-launch default browser in a split second
    # ?²æ­¢ Flask ??Debug Mode ä¸‹å???Reloader æ©Ÿåˆ¶?Ÿå??©æ¬¡?Œé??Ÿå…©?‹ç¶²??
    if not os.environ.get("WERKZEUG_RUN_MAIN"):
        Timer(1.0, launch_browser).start()
    app.run(port=5000, debug=True)


